import streamlit as st
import openpyxl
from datetime import datetime
import pandas as pd

st.set_page_config(page_title="Excel Data Processor", layout="centered")

st.title("Have there been any changes in site personnel? ")
st.write("Script does not work properly when PI changes.")
st.write("Date format should be DD.MM.YYYY")

st.write("Drag and Drop your Preperation list of the visit")

# -------------------------
# 1. FILE UPLOAD
# -------------------------

uploaded_file = st.file_uploader("Upload an Excel file (.xlsx)", type=["xlsx"])

if uploaded_file:
    try:
        workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
        sheet_names = workbook.sheetnames
    except Exception as e:
        st.error(f"Failed to load Excel file: {e}")
        st.stop()

    sheet_name = st.selectbox("Select Sheet:", sheet_names)
else:
    sheet_name = None

# -------------------------
# 2. COLUMN INPUTS
# -------------------------

st.subheader("Column Names")

beginn_header = st.text_input("Column name for 'Start Date':", "Beginn (Datum)")
ende_header = st.text_input("Column name for 'End Date':", "Ende (Datum)")
beteiligte_header = st.text_input("Column name for 'Participants':", "Beteiligte")
funktion_header = st.text_input("Column name for 'Function':", "Funktion")

# -------------------------
# 3. DATE INPUT
# -------------------------

user_date_input = st.text_input(
    "Enter date (Last Monitoring Visit / Initiation) (DD.MM.YYYY):"
)

# -------------------------
# 4. PROCESS BUTTON
# -------------------------

if st.button("Process Data"):
    if not uploaded_file or not sheet_name:
        st.error("Please upload a file and select a sheet.")
        st.stop()

    try:
        workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
        sheet = workbook[sheet_name]
    except Exception as e:
        st.error(f"Failed to load sheet: {e}")
        st.stop()

    # Convert sheet to pandas DataFrame to simplify work
    data = list(sheet.values)
    df = pd.DataFrame(data[1:], columns=data[0])

    # Validate column names
    required_cols = [beginn_header, beteiligte_header, funktion_header, ende_header]

    for col in required_cols:
        if col not in df.columns:
            st.error(f"Column '{col}' not found. Please check the header names.")
            st.stop()

    # Parse the user date
    try:
        user_date = datetime.strptime(user_date_input, "%d.%m.%Y")
    except ValueError:
        st.error("Invalid date format. Use DD.MM.YYYY.")
        st.stop()

    # Prepare output
    result_text = f"**Since the last monitoring visit on {user_date_input}, the following people joined the study group:**\n\n"

    # PROCESS BEGINN
    for _, row in df.iterrows():
        date_value = row[beginn_header]
        if pd.notna(date_value):
            try:
                excel_date = pd.to_datetime(date_value)
            except Exception:
                continue

            if excel_date > user_date:
                result_text += f"- **{row[beteiligte_header]}** ({row[funktion_header]}) on {excel_date.strftime('%d.%m.%Y')}\n"

    result_text += f"\n\n**Since the last monitoring visit on {user_date_input}, the following people left the study group:**\n\n"

    # PROCESS ENDE
    for _, row in df.iterrows():
        date_value = row[ende_header]
        if pd.notna(date_value):
            try:
                excel_date = pd.to_datetime(date_value)
            except Exception:
                continue

            if excel_date > user_date:
                result_text += f"- **{row[beteiligte_header]}** ({row[funktion_header]}) on {excel_date.strftime('%d.%m.%Y')}\n"

    # RESULT OUTPUT
    st.markdown(result_text)

