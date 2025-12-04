import openpyxl
from datetime import datetime
import os

def get_input(prompt, default=None):
    """Eingabe mit optionalem Default-Wert."""
    value = input(f"{prompt} [{default}]: ") if default else input(f"{prompt}: ")
    return value.strip() if value.strip() else default


def load_workbook_from_user():
    while True:
        file_path = input("Bitte den Pfad zur Excel-Datei eingeben: ").strip()

        if not os.path.isfile(file_path):
            print("❌ Datei wurde nicht gefunden. Bitte erneut versuchen.")
            continue

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            print("✔ Datei erfolgreich geladen.")
            return workbook, file_path
        except Exception as e:
            print(f"❌ Fehler beim Laden der Excel-Datei: {e}")


def choose_sheet(workbook):
    sheets = workbook.sheetnames
    print("\nVerfügbare Sheets:")
    for i, s in enumerate(sheets, 1):
        print(f"{i}: {s}")

    while True:
        choice = input("Bitte die Nummer des gewünschten Sheets eingeben: ").strip()
        if choice.isdigit() and 1 <= int(choice) <= len(sheets):
            return sheets[int(choice) - 1]
        print("❌ Ungültige Auswahl. Bitte erneut versuchen.")


def process_excel_data(workbook, sheet_name):
    sheet = workbook[sheet_name]

    # Spaltennamen abfragen
    beginn_header = get_input("Spaltenname für 'Start Date'", "Beginn (Datum)")
    ende_header = get_input("Spaltenname für 'End Date'", "Ende (Datum)")
    beteiligte_header = get_input("Spaltenname für 'Participants'", "Beteiligte")
    funktion_header = get_input("Spaltenname für 'Function'", "Funktion")

    # Datum abfragen
    while True:
        user_date_str = input("\nDatum der letzten Monitoring-Visite / Initiation (DD.MM.YYYY): ").strip()
        try:
            user_date = datetime.strptime(user_date_str, "%d.%m.%Y")
            break
        except ValueError:
            print("❌ Ungültiges Format. Bitte DD.MM.YYYY nutzen.")

    # Spalten-Indices finden
    column_headers = {"Beginn": None, "Beteiligte": None, "Funktion": None, "Ende": None}

    for cell in sheet[1]:
        if cell.value and isinstance(cell.value, str):
            if beginn_header in cell.value:
                column_headers["Beginn"] = cell.column
            elif beteiligte_header in cell.value:
                column_headers["Beteiligte"] = cell.column
            elif funktion_header in cell.value:
                column_headers["Funktion"] = cell.column
            elif ende_header in cell.value:
                column_headers["Ende"] = cell.column

    if not all(column_headers.values()):
        print("\n❌ Nicht alle benötigten Spalten wurden gefunden. Prüfe bitte die Spaltennamen.")
        return

    # Ausgabe vorbereiten
    result = []
    result.append(f"\nSeit der letzten MV am {user_date_str} sind folgende Personen NEU im Studienteam:\n")

    # Neue Teammitglieder
    for row in sheet.iter_rows(min_row=2, values_only=False):
        beginn = row[column_headers["Beginn"] - 1].value
        beteiligte = row[column_headers["Beteiligte"] - 1].value
        funktion = row[column_headers["Funktion"] - 1].value

        if beginn:
            excel_date = parse_excel_date(beginn)
            if excel_date and excel_date > user_date:
                result.append(f"- {beteiligte} ({funktion}) seit {excel_date.strftime('%d.%m.%Y')}")

    result.append(f"\nSeit der letzten MV am {user_date_str} haben folgende Personen das Studienteam VERLASSEN:\n")

    # Ausgestiegene Teammitglieder
    for row in sheet.iter_rows(min_row=2, values_only=False):
        ende = row[column_headers["Ende"] - 1].value
        beteiligte = row[column_headers["Beteiligte"] - 1].value
        funktion = row[column_headers["Funktion"] - 1].value

        if ende:
            excel_date = parse_excel_date(ende)
            if excel_date and excel_date > user_date:
                result.append(f"- {beteiligte} ({funktion}) am {excel_date.strftime('%d.%m.%Y')}")

    # Ausgabe
    print("\n".join(result))


def parse_excel_date(value):
    """Hilfsfunktion für flexible Datumserkennung."""
    if isinstance(value, datetime):
        return value
    try:
        return datetime.strptime(str(value), "%Y-%m-%d %H:%M:%S")
    except:
        try:
            return datetime.strptime(str(value), "%Y-%m-%d")
        except:
            return None


def main():
    print("\n=== Excel Data Processor (Terminal Version) ===\n")

    workbook, file_path = load_workbook_from_user()
    sheet_name = choose_sheet(workbook)

    print(f"\n✔ Ausgewähltes Sheet: {sheet_name}")

    process_excel_data(workbook, sheet_name)

    print("\n=== Fertig ===\n")


if __name__ == "__main__":
    main()
